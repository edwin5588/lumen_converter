import tempfile
import os
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import conversion_script

st.title("Lumen Converter")

# --- Well selector grid ---
st.subheader("Well Selection")
st.caption("Use row/column selectors for bulk selection, or click individual cells. Leave all unchecked to include everything.")

rows = list("ABCDEFGH")
cols = [f"{j:02d}" for j in range(1, 13)]

if "well_grid" not in st.session_state:
    st.session_state.well_grid = pd.DataFrame(False, index=rows, columns=cols)

# Bulk selection controls
sel_rows = st.multiselect("Rows", rows, placeholder="Select rows…", label_visibility="collapsed")
sel_cols = st.multiselect("Columns", cols, placeholder="Select columns…", label_visibility="collapsed")

b1, b2, b3 = st.columns([1, 1, 1])
if b1.button("Apply selection", use_container_width=True):
    target_rows = sel_rows or rows
    target_cols = sel_cols or cols
    for r in target_rows:
        for c in target_cols:
            st.session_state.well_grid.loc[r, c] = True
    st.rerun()
if b2.button("Select all", use_container_width=True):
    st.session_state.well_grid = pd.DataFrame(True, index=rows, columns=cols)
    st.rerun()
if b3.button("Clear all", use_container_width=True):
    st.session_state.well_grid = pd.DataFrame(False, index=rows, columns=cols)
    st.rerun()

# Grid display — narrow fixed-px columns so all 12 fit without horizontal scroll
st.markdown("""<style>
[data-testid="stDataEditor"] > div { overflow-x: hidden !important; }
</style>""", unsafe_allow_html=True)
col_cfg = {"_index": st.column_config.Column(width=28)}
col_cfg.update({c: st.column_config.CheckboxColumn(c, width=46) for c in cols})
edited = st.data_editor(
    st.session_state.well_grid,
    column_config=col_cfg,
    use_container_width=True,
    height=334,
    key="well_grid_editor",
)
st.session_state.well_grid = edited

selected_wells = {
    f"{r}{c}"
    for r in edited.index
    for c in edited.columns
    if edited.loc[r, c]
}

if selected_wells:
    st.caption(f"Selected ({len(selected_wells)}): " + ", ".join(sorted(selected_wells)))
else:
    st.caption("No wells selected — all wells will be included.")

# --- File upload ---
st.subheader("Upload Files")
st.write("Upload plate reader `.xls` or `.xlsx` files to convert them into formatted plates.")

uploaded_files = st.file_uploader("Choose files", type=["xls", "xlsx"], accept_multiple_files=True)


def apply_well_mask(xlsx_path, wells):
    if not wells:
        return

    wb = load_workbook(xlsx_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=False))
        col_map = {}

        for row in all_rows:
            first_val = row[0].value

            if first_val is None or str(first_val).strip() == "":
                if len(row) > 1 and row[1].value is not None:
                    second = row[1].value
                    second_str = f"{second:02d}" if isinstance(second, int) else str(second).strip()
                    if second_str == "01":
                        col_map = {}
                        for cell in row[1:]:
                            if cell.value is not None:
                                v = cell.value
                                col_map[cell.column] = f"{v:02d}" if isinstance(v, int) else str(v).strip()
                continue

            row_label = str(first_val).strip()
            if row_label in set("ABCDEFGH") and col_map:
                for cell in row[1:]:
                    col_str = col_map.get(cell.column)
                    if col_str and f"{row_label}{col_str}" not in wells:
                        cell.value = 0

    wb.save(xlsx_path)


for uploaded in uploaded_files:
    suffix = os.path.splitext(uploaded.name)[1]
    out_path = None
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_in:
        tmp_in.write(uploaded.read())
        tmp_in_path = tmp_in.name

    try:
        with st.spinner(f"Converting {uploaded.name}..."):
            out_path = conversion_script.convert(tmp_in_path)
            apply_well_mask(out_path, selected_wells)

        with open(out_path, "rb") as f:
            out_bytes = f.read()

        out_name = os.path.splitext(uploaded.name)[0] + "_CONVERTED_plates.xlsx"
        st.success(f"{uploaded.name} converted.")
        st.download_button(
            label=f"Download {out_name}",
            data=out_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=uploaded.name,
        )
    except Exception as e:
        st.error(f"{uploaded.name} failed: {e}")
    finally:
        os.unlink(tmp_in_path)
        if out_path and os.path.exists(out_path):
            os.unlink(out_path)
